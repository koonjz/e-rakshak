from fastapi import FastAPI, HTTPException, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import time
import asyncio
import os
import sys
from typing import Dict, Any, List

# Ensure parent directory is in sys.path so we can import the ml module
import os
from dotenv import load_dotenv
current_dir = os.path.dirname(os.path.abspath(__file__))
dotenv_path = os.path.abspath(os.path.join(current_dir, "..", ".env"))
load_dotenv(dotenv_path=dotenv_path)

sys.path.append(os.path.abspath(os.path.join(current_dir, "..")))

from crawler.mock import MockCrawler
from crawler.social_stubs import YouTubeCrawler, InstagramCrawler, FacebookCrawler, TelegramCrawler, TwitterCrawler
from ml.classifier import MultilingualThreatClassifier
from ml.image_analyzer import analyze_image

app = FastAPI(title="Social Threat Analyzer API")

# Configure CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for dev stack verification
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global Classifier, Crawler, and In-Memory Post Store State
class AppState:
    def __init__(self):
        self.task: asyncio.Task = None
        self.queue: asyncio.Queue = asyncio.Queue(maxsize=2000)
        self.crawler = MockCrawler(lookback_days=30)
        self.active = False
        self.mode = "mock"  # Current active crawler mode: "mock" or "youtube"
        # Instantiate the TF-IDF hybrid classifier once
        self.classifier = MultilingualThreatClassifier()
        # In-memory database of all classified posts (pre-populated & real-time)
        self.posts_db: List[Dict[str, Any]] = []
        
        # Prepopulate db with mock posts on startup
        self.prepopulate_db()

    def prepopulate_db(self):
        print("Pre-populating in-memory database...")
        posts = self.crawler.posts
        for post in posts:
            try:
                # Classify and enrich at startup
                classification = self.classifier.predict(post["text"])
                post["language"] = classification["language"]
                post["threat_category"] = classification["threat_category"]
                post["classification_meta"] = classification
                self.posts_db.append(post)
            except Exception as e:
                print(f"Error pre-populating post {post.get('id')}: {e}")
        print(f"In-memory database pre-populated with {len(self.posts_db)} posts.")

    def sync_to_graph_db(self):
        """Syncs current in-memory database posts and detected coordination clusters to Neo4j graph database."""
        try:
            from analytics.graph_db import graph_manager
            from analytics.coordination import detect_coordinated_behavior
            if graph_manager.available:
                clusters = detect_coordinated_behavior(self.posts_db)
                graph_manager.sync_posts_and_coordination(self.posts_db, clusters)
                print("Neo4j Graph Database successfully synchronized.")
            else:
                print(f"Neo4j synchronization bypassed: {graph_manager.error_message}")
        except Exception as e:
            print(f"Failed to sync to Neo4j: {e}")

state = AppState()

@app.on_event("startup")
async def startup_event():
    print("FastAPI server startup complete. Triggering background Neo4j sync...")
    try:
        asyncio.create_task(asyncio.to_thread(state.sync_to_graph_db))
    except Exception as e:
        print(f"Error starting background Neo4j sync task: {e}")

# Pydantic request models
class ClassifyRequest(BaseModel):
    text: str

async def crawler_worker(keywords: List[str] = None):
    """
    Background worker that streams posts from the crawler and pushes them to the queue.
    """
    try:
        print(f"Crawler worker task started with keywords: {keywords}")
        state.last_crawler_error = None
        async for post in state.crawler.stream_posts(keywords=keywords):
            if isinstance(post, dict) and post.get("status") == "error":
                print(f"Crawler worker received error payload: {post}")
                state.last_crawler_error = post
                state.active = False
                break
                
            # If the queue gets completely full, discard the oldest to avoid memory leaks
            if state.queue.full():
                try:
                    state.queue.get_nowait()
                except asyncio.QueueEmpty:
                    pass
            
            # Enrich dynamic stream posts with model classification in real-time
            try:
                classification = state.classifier.predict(post["text"])
                post["language"] = classification["language"]
                post["threat_category"] = classification["threat_category"]
                post["classification_meta"] = classification
            except Exception as ml_err:
                print(f"Error classifying live stream post: {ml_err}")
                
            # Add to in-memory store for historical analytics
            state.posts_db.append(post)
            
            # Sync to Neo4j database asynchronously in a worker thread
            try:
                if state.active:
                    asyncio.create_task(asyncio.to_thread(state.sync_to_graph_db))
            except Exception as sync_err:
                print(f"Error launching background Neo4j sync task: {sync_err}")
            
            await state.queue.put(post)
    except asyncio.CancelledError:
        print("Crawler worker task cancelled by manager.")
    except Exception as e:
        print(f"Error in crawler worker task: {e}")
    finally:
        state.active = False
        state.task = None
        print("Crawler worker task shut down.")

@app.get("/api/health")
def health_check():
    return {
        "status": "healthy",
        "timestamp": time.time(),
        "service": "social-threat-analyzer-backend",
        "version": "0.1.0",
        "gemini_model": os.getenv("GEMINI_MODEL", "gemini-3.5-flash")
    }

@app.post("/api/classify")
def classify_text(request: ClassifyRequest):
    """
    Classify a single piece of social media text for language, sentiment, threat category, and confidence.
    """
    try:
        return state.classifier.predict(request.text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Classifier error: {str(e)}")

@app.get("/api/trends")
def get_trends(threat_category: str = None, interval: str = "day"):
    """
    Query real-time and historical trends including keywords, spikes, and geo-breakdowns.
    """
    try:
        from analytics.trends import compute_trends
        return compute_trends(state.posts_db, threat_category=threat_category, interval=interval)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to compute trends: {str(e)}")

@app.get("/api/coordination")
def get_coordination():
    """
    Detect likely coordinated bot-like amplification clusters.
    """
    try:
        from analytics.coordination import detect_coordinated_behavior
        return detect_coordinated_behavior(state.posts_db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to detect coordinated behavior: {str(e)}")

@app.get("/api/incidents")
def get_incidents():
    """
    Get persistent incident records generated from incitement to violence or coordination clusters.
    """
    try:
        from analytics.incidents import get_all_incidents
        return get_all_incidents(state.posts_db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to compile incidents: {str(e)}")

@app.get("/api/incidents/export")
def export_incidents_excel(n: int = 10):
    """
    Export the top N threat incidents sorted by reach/follower count descending to an Excel file.
    """
    try:
        from analytics.incidents import get_all_incidents
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        
        # Get all incidents
        incidents = get_all_incidents(state.posts_db)
        
        # Helper to get reach (max follower count)
        def get_incident_follower_count(inc: dict) -> float:
            related_posts = inc.get("related_posts", [])
            if not related_posts:
                return float('-inf')
            counts = []
            for p in related_posts:
                val = p.get("user_profile", {}).get("follower_count")
                if val is not None and isinstance(val, (int, float)):
                    counts.append(val)
            if not counts:
                return float('-inf')
            return max(counts)
            
        # Stable sort matching frontend:
        # 1. Sort by incident_id ascending
        incidents.sort(key=lambda x: x.get("incident_id", ""))
        # 2. Sort by timestamp descending
        incidents.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
        # 3. Sort by reach (max follower_count) descending
        incidents.sort(key=get_incident_follower_count, reverse=True)
        
        # Slice top N
        top_n = incidents[:n]
        
        # Build Excel sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Threat Incidents"
        
        # Headers
        headers = [
            "Incident ID", "Timestamp", "Threat Category", "Severity", 
            "Platform", "Account/Username", "Follower Count", 
            "Geo/City", "Post Text Snippet", "Suspicion Score", "Escalation Template"
        ]
        ws.append(headers)
        
        # Styling header: bold
        bold_font = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = bold_font
            
        # Freeze headers row
        ws.freeze_panes = "A2"
        
        # Append data rows
        for inc in top_n:
            inc_id = inc.get("incident_id", "")
            timestamp = inc.get("timestamp", "")
            category = inc.get("threat_category", "")
            severity = inc.get("severity", "")
            
            related_posts = inc.get("related_posts", [])
            platforms = ", ".join(sorted(list(set(p.get("platform", "X") for p in related_posts))))
            usernames = ", ".join(p.get("username", "unknown") for p in related_posts)
            
            reach = get_incident_follower_count(inc)
            reach_str = str(reach) if reach != float('-inf') else "N/A"
            
            geo = inc.get("affected_geo", "")
            
            # Post Text Snippet: text of the first post
            text_snippet = related_posts[0].get("text", "") if related_posts else ""
            
            suspicion = inc.get("suspicion_score", "")
            suspicion_str = f"{suspicion}%" if suspicion != "" else "N/A"
            
            template = inc.get("suggested_escalation_template", "")
            
            ws.append([
                inc_id, timestamp, category, severity,
                platforms, usernames, reach_str,
                geo, text_snippet, suspicion_str, template
            ])
            
        # Auto-fit column widths (cap at 50)
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                lines = val_str.split("\n")
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        # Write to byte stream
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=top_{n}_incidents_export.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export incidents: {str(e)}")

@app.get("/api/incidents/{incident_id}/pdf")
def get_incident_pdf(incident_id: str):
    """
    Generate and download a styled PDF report for a specific threat incident.
    """
    try:
        from analytics.incidents import get_all_incidents
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        # Get all incidents
        incidents = get_all_incidents(state.posts_db)
        incident = next((inc for inc in incidents if inc.get("incident_id") == incident_id), None)
        
        if not incident:
            raise HTTPException(status_code=404, detail=f"Incident {incident_id} not found")
            
        # Build PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=40, leftMargin=40,
                                topMargin=40, bottomMargin=40)
                                
        styles = getSampleStyleSheet()
        
        # Custom styles to prevent overlap and support nice layout
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#991b1b'), # Rose-800
            spaceAfter=15
        )
        
        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=16,
            textColor=colors.HexColor('#1f2937'), # Zinc-800
            spaceBefore=15,
            spaceAfter=8
        )
        
        normal_style = ParagraphStyle(
            'ReportNormal',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#374151') # Zinc-700
        )
        
        meta_label_style = ParagraphStyle(
            'MetaLabel',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#111827')
        )
        
        code_style = ParagraphStyle(
            'ReportCode',
            parent=styles['Code'],
            fontName='Courier',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#111827')
        )
        
        story = []
        
        # Title
        story.append(Paragraph(f"THREAT INCIDENT REPORT: {incident_id}", title_style))
        story.append(Spacer(1, 10))
        
        # Metadata Table
        meta_data = [
            [Paragraph("Severity:", meta_label_style), Paragraph(incident.get("severity", ""), normal_style),
             Paragraph("Affected Geo:", meta_label_style), Paragraph(incident.get("affected_geo", ""), normal_style)],
            [Paragraph("Category:", meta_label_style), Paragraph(incident.get("threat_category", ""), normal_style),
             Paragraph("Timestamp:", meta_label_style), Paragraph(incident.get("timestamp", ""), normal_style)]
        ]
        
        meta_table = Table(meta_data, colWidths=[80, 180, 80, 180])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f9fafb')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e5e7eb')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#f3f4f6')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 15))
        
        # Incident Summary
        story.append(Paragraph("Incident Summary", section_heading))
        story.append(Paragraph(incident.get("summary", ""), normal_style))
        story.append(Spacer(1, 15))
        
        # Related Posts
        story.append(Paragraph("Matched Social Media Source Posts", section_heading))
        related_posts = incident.get("related_posts", [])
        
        for idx, p in enumerate(related_posts, 1):
            post_intro = f"<b>Post #{idx} ({p.get('platform', 'X')})</b> - User: <b>@{p.get('username')}</b> | Timestamp: {p.get('timestamp')}"
            story.append(Paragraph(post_intro, normal_style))
            
            # Post Text boxed
            post_text_table = Table([[Paragraph(p.get("text", ""), normal_style)]], colWidths=[520])
            post_text_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f3f4f6')),
                ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('LEFTPADDING', (0,0), (-1,-1), 8),
                ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ]))
            story.append(Spacer(1, 4))
            story.append(post_text_table)
            story.append(Spacer(1, 10))
            
        story.append(Spacer(1, 5))
        
        # Suggested Escalation Template
        story.append(Paragraph("Suggested Duty Officer Escalation Template", section_heading))
        
        template_text = incident.get("suggested_escalation_template", "")
        template_paragraphs = [Paragraph(line, code_style) for line in template_text.split('\n')]
        
        template_cell = []
        for tp in template_paragraphs:
            template_cell.append(tp)
            
        template_table = Table([[template_cell]], colWidths=[520])
        template_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#fffbeb')), # Light amber
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#fef3c7')),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(template_table)
        story.append(Spacer(1, 20))
        
        # Footer notice
        footer_style = ParagraphStyle(
            'ReportFooter',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#9ca3af'),
            alignment=1 # Center
        )
        story.append(Paragraph(f"Report generated programmatically via Social Threat Analyzer Console on {time.strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        doc.build(story)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=incident_report_{incident_id}.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")

@app.get("/api/network-graph")
def get_network_graph():
    """
    Get accounts and coordination relationships formatted for force-directed graph.
    """
    try:
        from analytics.graph_db import graph_manager
        res = graph_manager.get_network_graph()
        
        # Fallback to computing graph from in-memory coordination clusters if Neo4j is offline
        if not graph_manager.available:
            from analytics.coordination import detect_coordinated_behavior
            clusters = detect_coordinated_behavior(state.posts_db)
            
            nodes = []
            edges = []
            node_map = {}
            
            for cluster in clusters:
                member_accounts = cluster.get("member_accounts", [])
                heuristics = cluster.get("heuristics", [])
                heuristic_label = ", ".join(heuristics) if heuristics else "Coordinated"
                suspicion_score = cluster.get("suspicion_score", 50)
                
                for username in member_accounts:
                    matched_posts = [p for p in state.posts_db if p.get("username") == username]
                    platform = matched_posts[0].get("platform", "Mock") if matched_posts else "Mock"
                    post_count = len(matched_posts)
                    
                    if username not in node_map:
                        node_map[username] = {
                            "id": username,
                            "label": username,
                            "platform": platform,
                            "post_count": post_count,
                            "suspicion": suspicion_score
                        }
                    else:
                        node_map[username]["suspicion"] = max(node_map[username]["suspicion"], suspicion_score)
                        node_map[username]["post_count"] = max(node_map[username]["post_count"], post_count)
                        
                n_members = len(member_accounts)
                for i in range(n_members):
                    for j in range(i + 1, n_members):
                        u1, u2 = sorted([member_accounts[i], member_accounts[j]])
                        edges.append({
                            "from": u1,
                            "to": u2,
                            "heuristic": heuristic_label,
                            "suspicion_score": suspicion_score
                        })
                        
            # Include other active posters as neutral background nodes
            for p in state.posts_db:
                username = p.get("username")
                if username and username not in node_map:
                    node_map[username] = {
                        "id": username,
                        "label": username,
                        "platform": p.get("platform", "Mock"),
                        "post_count": 1,
                        "suspicion": 0
                    }
                    
            res = {
                "status": "fallback_success",
                "nodes": list(node_map.values()),
                "edges": edges
            }
            
        return {
            "neo4j_available": graph_manager.available,
            "status": res["status"],
            "nodes": res["nodes"],
            "edges": res["edges"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch network graph: {str(e)}")

async def stop_crawler_helper():
    if state.active and state.task:
        state.task.cancel()
        for _ in range(10):
            if state.task is None or state.task.done():
                break
            await asyncio.sleep(0.1)
        state.active = False
        state.task = None

@app.post("/api/crawler/start")
async def start_crawler(lookback_days: int = 7):
    if state.active:
        if state.mode == "mock":
            return {
                "status": "already_running",
                "message": "Crawler background worker is already active.",
                "queue_size": state.queue.qsize()
            }
        else:
            await stop_crawler_helper()
    
    state.crawler = MockCrawler(lookback_days=lookback_days)
    state.mode = "mock"
    state.active = True
    state.task = asyncio.create_task(crawler_worker())
    return {
        "status": "started",
        "message": f"Mock Crawler started successfully with lookback window of {lookback_days} days."
    }

@app.post("/api/crawler/start-live")
async def start_live_crawler(platform: str = "youtube", keywords: str = "Gujarat", lookback_days: int = 7):
    # Enforce access checks for Meta platforms
    meta_access_token = os.getenv("META_ACCESS_TOKEN")
    
    if platform.lower() in ("instagram", "facebook"):
        if not meta_access_token:
            return {
                "status": "error",
                "message": "No credentials configured: META_ACCESS_TOKEN is missing."
            }
            
    # Enforce access checks for Twitter
    twitter_bearer_token = os.getenv("TWITTER_BEARER_TOKEN")
    if platform.lower() in ("twitter", "x"):
        if not twitter_bearer_token:
            return {
                "status": "error",
                "message": "No credentials configured: TWITTER_BEARER_TOKEN is missing. Ingestion is blocked by X paid-API requirement."
            }
            
    # Enforce access checks for Telegram
    if platform.lower() == "telegram":
        telegram_api_id = os.getenv("TELEGRAM_API_ID")
        telegram_api_hash = os.getenv("TELEGRAM_API_HASH")
        current_dir = os.path.dirname(os.path.abspath(__file__))
        session_file = os.path.abspath(os.path.join(current_dir, "..", "telegram_session.session"))
        session_exists = os.path.exists(session_file)
        
        if not telegram_api_id or not telegram_api_hash or not session_exists:
            return {
                "status": "pending_auth",
                "message": "Telegram API credentials not configured or session not authenticated. Run login flow first."
            }
            
    # If active, stop the current worker before switching modes
    await stop_crawler_helper()
    
    warning = None
    if platform.lower() in ("twitter", "x") and lookback_days > 7:
        warning = "X (Twitter) recent search endpoint limits lookback to the last 7 days."
    elif platform.lower() in ("instagram", "facebook") and lookback_days > 7:
        warning = "Meta Graph API recent hashtag/feed endpoints typically limit lookback to the last 7 days."
    
    if platform.lower() == "instagram":
        state.crawler = InstagramCrawler(lookback_days=lookback_days)
        state.mode = "instagram"
    elif platform.lower() == "facebook":
        state.crawler = FacebookCrawler(lookback_days=lookback_days)
        state.mode = "facebook"
    elif platform.lower() in ("twitter", "x"):
        state.crawler = TwitterCrawler(lookback_days=lookback_days)
        state.mode = "twitter"
    elif platform.lower() == "telegram":
        state.crawler = TelegramCrawler(lookback_days=lookback_days)
        state.mode = "telegram"
    else:
        state.crawler = YouTubeCrawler(lookback_days=lookback_days)
        state.mode = "youtube"
        
    state.active = True
    
    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]
    state.task = asyncio.create_task(crawler_worker(keywords=kw_list))
    res_payload = {
        "status": "started",
        "message": f"Live {platform.upper()} Crawler started for keywords: {keywords}."
    }
    if warning:
        res_payload["warning"] = warning
    return res_payload

@app.post("/api/crawler/stop")
async def stop_crawler():
    if not state.active:
        return {
            "status": "not_running",
            "message": "Crawler background worker is not currently active."
        }
    
    await stop_crawler_helper()
    return {
        "status": "stopped",
        "message": "Crawler background worker stopped successfully."
    }

@app.get("/api/crawler/status")
def get_crawler_status():
    crawler_name = type(state.crawler).__name__
    if crawler_name == "MockCrawler":
        dataset_path = state.crawler.data_path
        dataset_exists = os.path.exists(dataset_path)
    else:
        dataset_path = f"{crawler_name} API Connection"
        dataset_exists = True
        
    current_dir = os.path.dirname(os.path.abspath(__file__))
    session_file = os.path.abspath(os.path.join(current_dir, "..", "telegram_session.session"))
    session_exists = os.path.exists(session_file)
        
    return {
        "active": state.active,
        "mode": state.mode,
        "queue_size": state.queue.qsize(),
        "dataset_path": dataset_path,
        "dataset_exists": dataset_exists,
        "youtube_key_loaded": bool(os.getenv("YOUTUBE_API_KEY")),
        "meta_token_loaded": bool(os.getenv("META_ACCESS_TOKEN")),
        "twitter_auth_loaded": bool(os.getenv("TWITTER_BEARER_TOKEN")),
        "telegram_auth_loaded": bool(os.getenv("TELEGRAM_API_ID") and os.getenv("TELEGRAM_API_HASH") and session_exists),
        "last_error": getattr(state, "last_crawler_error", None)
    }

@app.get("/api/crawler/posts", response_model=List[Dict[str, Any]])
def get_crawled_posts(limit: int = 50):
    posts = []
    # Drain posts from queue up to the specified limit
    qsize = state.queue.qsize()
    drain_count = min(limit, qsize)
    
    for _ in range(drain_count):
        try:
            post = state.queue.get_nowait()
            posts.append(post)
        except asyncio.QueueEmpty:
            break
            
    return posts

@app.post("/api/analyze-image")
async def analyze_image_route(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        result = analyze_image(contents)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Image analysis failed: {str(e)}")

class AssistantQueryRequest(BaseModel):
    question: str

@app.post("/api/assistant/query")
def assistant_query(request: AssistantQueryRequest):
    try:
        from assistant.query_router import answer_question
        res = answer_question(request.question, state.posts_db)
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Assistant query failed: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
