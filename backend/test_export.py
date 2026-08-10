import urllib.request
import io
import sys
import openpyxl

# Enforce stdout to UTF-8
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

def test_export(n):
    url = f"http://127.0.0.1:8000/api/incidents/export?n={n}"
    print(f"Requesting: {url}")
    
    try:
        response = urllib.request.urlopen(url)
        content_type = response.headers.get("Content-Type")
        content_disposition = response.headers.get("Content-Disposition")
        
        print(f"Status Code: {response.status}")
        print(f"Content-Type: {content_type}")
        print(f"Content-Disposition: {content_disposition}")
        
        # Read into bytes
        data = response.read()
        print(f"Downloaded file size: {len(data)} bytes")
        
        # Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(data))
        print("Workbook loaded successfully!")
        
        ws = wb.active
        print(f"Active Sheet Title: {ws.title}")
        
        # Check header
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            "Incident ID", "Timestamp", "Threat Category", "Severity", 
            "Platform", "Account/Username", "Follower Count", 
            "Geo/City", "Post Text Snippet", "Suspicion Score", "Escalation Template"
        ]
        assert headers == expected_headers, f"Headers mismatch: {headers}"
        print("Headers check passed!")
        
        # Check bold formatting on headers
        for cell in ws[1]:
            assert cell.font.bold is True, f"Cell {cell.column} is not bold"
        print("Header bold styling check passed!")
        
        # Print data rows
        rows = list(ws.iter_rows(values_only=True))
        print(f"Total rows in sheet (including header): {len(rows)}")
        
        # Verify sorting by follower count descending
        follower_counts = []
        for r_idx in range(1, len(rows)):
            val = rows[r_idx][6] # 7th column: Follower Count
            if val != "N/A":
                follower_counts.append(int(val))
        
        print("Follower counts in order:", follower_counts)
        # Check if list is sorted descending
        is_sorted = all(follower_counts[i] >= follower_counts[i+1] for i in range(len(follower_counts)-1))
        assert is_sorted, "Data is not sorted by follower count descending!"
        print("Data sorting check passed!")
        
        print(f"=== Top 3 data rows (N={n}) ===")
        for i in range(1, min(len(rows), 4)):
            print(f"Row {i}: {rows[i][:8]} (Snippet: {rows[i][8][:30]}...)")
            
        print("-" * 50)
        
    except Exception as e:
        print(f"Verification failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    print("Seeding some mock crawler runs to populate database first...")
    # Trigger a mock crawler run brief sync to populate posts and incidents
    try:
        urllib.request.urlopen("http://127.0.0.1:8000/api/crawler/start", data=b"")
        import time
        time.sleep(3) # Wait 3 seconds for mock sync to ingest posts
        urllib.request.urlopen("http://127.0.0.1:8000/api/crawler/stop", data=b"")
        print("Mock seeding complete.")
    except Exception as ex:
        print(f"Seeding note: {ex}")
        
    print("\n--- Testing N=5 ---")
    test_export(5)
    
    print("\n--- Testing N=50 ---")
    test_export(50)
